from __future__ import annotations

import logging
import re
from functools import lru_cache
from pathlib import Path
from typing import Any

from langchain_community.document_loaders import (
    Docx2txtLoader,
    PyPDFLoader,
    TextLoader,
    UnstructuredWordDocumentLoader,
)
from langchain_community.vectorstores import FAISS
from langchain_core.documents import Document
from langchain_openai import OpenAIEmbeddings
from langchain_text_splitters import RecursiveCharacterTextSplitter
from werkzeug.datastructures import FileStorage
from werkzeug.utils import secure_filename

from enterpriseRAG.config.settings import SUPPORTED_EXTENSIONS, UPLOADS_DIR, get_settings
from enterpriseRAG.services.observability import ObservabilityService

logger = logging.getLogger(__name__)


def _chunk_section_label(content: str, fallback: str = "Section") -> str:
    """Derive section title; skip URLs, watermarks, and page-number-only lines."""
    skip = re.compile(
        r"^(https?://|www\.|\d+\s*$|dailydose|\.com\b|page \d)",
        re.IGNORECASE,
    )
    for line in content.splitlines():
        stripped = line.strip()
        if not stripped or len(stripped) <= 3:
            continue
        if skip.search(stripped):
            continue
        if re.match(r"^[\d\s\.\-]+$", stripped):
            continue
        return stripped[:120]
    return fallback


def _load_excel(file_path: Path) -> list[Document]:
    suffix = file_path.suffix.lower()
    docs: list[Document] = []
    if suffix == ".xlsx":
        import openpyxl

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(cells):
                    rows.append(" | ".join(cells))
            if rows:
                docs.append(
                    Document(
                        page_content="\n".join(rows),
                        metadata={"section": sheet_name, "page": 1, "source": str(file_path)},
                    )
                )
        wb.close()
    elif suffix == ".xls":
        import xlrd

        wb = xlrd.open_workbook(str(file_path))
        for sheet in wb.sheets():
            rows = []
            for rx in range(sheet.nrows):
                cells = [str(sheet.cell_value(rx, cx)) for cx in range(sheet.ncols)]
                if any(cells):
                    rows.append(" | ".join(cells))
            if rows:
                docs.append(
                    Document(
                        page_content="\n".join(rows),
                        metadata={"section": sheet.name, "page": 1, "source": str(file_path)},
                    )
                )
    return docs


class DocumentService:
    """Document loading, chunking, and index management."""

    def __init__(self, observability: ObservabilityService | None = None) -> None:
        self.settings = get_settings()
        self.obs = observability
        self._ensure_dirs()

    def _ensure_dirs(self) -> None:
        UPLOADS_DIR.mkdir(parents=True, exist_ok=True)

    def clear_uploads(self) -> None:
        for item in UPLOADS_DIR.iterdir():
            if item.is_file():
                item.unlink()
        self._invalidate_caches()

    def _invalidate_caches(self) -> None:
        self.get_indexes.cache_clear()

    @staticmethod
    def get_uploaded_file() -> Path | None:
        files = [
            f
            for f in UPLOADS_DIR.iterdir()
            if f.is_file() and f.name != ".gitkeep" and not f.name.startswith(".")
        ]
        return files[0] if files else None

    def save_upload(self, file_storage: FileStorage) -> Path:
        if not file_storage or not file_storage.filename:
            raise ValueError("No file was uploaded.")
        suffix = Path(file_storage.filename).suffix.lower()
        if suffix not in SUPPORTED_EXTENSIONS:
            supported = ", ".join(sorted(SUPPORTED_EXTENSIONS))
            raise ValueError(f"Unsupported file type. Allowed: {supported}")

        self.clear_uploads()
        safe_name = secure_filename(file_storage.filename)
        dest = UPLOADS_DIR / safe_name
        file_storage.save(dest)
        self._invalidate_caches()
        logger.info("Saved enterprise RAG upload: %s", dest.name)
        return dest

    def load_document(self, file_path: Path) -> list[Document]:
        suffix = file_path.suffix.lower()
        if suffix == ".pdf":
            docs = PyPDFLoader(str(file_path)).load()
        elif suffix == ".txt":
            docs = TextLoader(str(file_path), encoding="utf-8").load()
        elif suffix == ".docx":
            docs = Docx2txtLoader(str(file_path)).load()
        elif suffix == ".doc":
            docs = UnstructuredWordDocumentLoader(str(file_path)).load()
        elif suffix in (".xlsx", ".xls"):
            docs = _load_excel(file_path)
        else:
            raise ValueError(f"Unsupported format: {suffix}")

        for doc in docs:
            page = int(doc.metadata.get("page", 0)) + 1
            doc.metadata["page"] = page
            doc.metadata["source"] = str(file_path)
            doc.metadata.setdefault(
                "section", _chunk_section_label(doc.page_content, file_path.stem)
            )
        return docs

    def _split_documents(self, docs: list[Document]) -> tuple[list[Document], list[Document]]:
        child_splitter = RecursiveCharacterTextSplitter(
            chunk_size=self.settings.chunk_size,
            chunk_overlap=self.settings.chunk_overlap,
        )
        parent_splitter = RecursiveCharacterTextSplitter(
            chunk_size=self.settings.parent_chunk_size,
            chunk_overlap=self.settings.chunk_overlap,
        )
        children = child_splitter.split_documents(docs)
        parents = parent_splitter.split_documents(docs)

        for i, chunk in enumerate(children):
            chunk.metadata["chunk_id"] = f"child_{i}"
            chunk.metadata["section"] = _chunk_section_label(
                chunk.page_content, chunk.metadata.get("section", "Section")
            )

        for i, chunk in enumerate(parents):
            chunk.metadata["chunk_id"] = f"parent_{i}"
            chunk.metadata["is_parent"] = True

        return children, parents

    @lru_cache(maxsize=4)
    def get_indexes(self, file_path_str: str, mtime_ns: int) -> dict[str, Any]:
        """Build FAISS vector index and BM25 corpus from uploaded document."""
        file_path = Path(file_path_str)
        docs = self.load_document(file_path)
        children, parents = self._split_documents(docs)

        embeddings = OpenAIEmbeddings(
            model=self.settings.embedding_model,
            openai_api_key=self.settings.openai_api_key,
        )
        vector_store = FAISS.from_documents(children, embeddings)

        child_texts = [d.page_content for d in children]
        child_metas = [d.metadata for d in children]

        return {
            "vector_store": vector_store,
            "children": children,
            "parents": parents,
            "child_texts": child_texts,
            "child_metas": child_metas,
            "document_name": file_path.name,
            "file_path": str(file_path),
        }

    def get_active_indexes(self) -> dict[str, Any]:
        uploaded = self.get_uploaded_file()
        if not uploaded:
            raise FileNotFoundError(
                "No document uploaded. Please upload a document first."
            )
        mtime = uploaded.stat().st_mtime_ns
        return self.get_indexes(str(uploaded), mtime)

    def preview_text(self, max_chars: int = 50000) -> str:
        uploaded = self.get_uploaded_file()
        if not uploaded:
            raise FileNotFoundError("No document uploaded.")
        docs = self.load_document(uploaded)
        text = "\n\n".join(d.page_content for d in docs)
        return text[:max_chars]

    def document_stats(self) -> dict[str, Any]:
        indexes = self.get_active_indexes()
        children = indexes["children"]
        parents = indexes["parents"]
        pages = {c.metadata.get("page", 0) for c in children}
        sections = {c.metadata.get("section", "") for c in children}
        return {
            "document_name": indexes["document_name"],
            "child_chunks": len(children),
            "parent_chunks": len(parents),
            "unique_pages": len(pages),
            "unique_sections": len(sections),
            "total_characters": sum(len(c.page_content) for c in children),
        }

    @staticmethod
    def chunk_to_dict(doc: Document, score: float = 0.0) -> dict[str, Any]:
        return {
            "content": doc.page_content,
            "section": doc.metadata.get("section", "Unknown"),
            "page": doc.metadata.get("page", 0),
            "chunk_id": doc.metadata.get("chunk_id", ""),
            "source": doc.metadata.get("source", ""),
            "score": round(score, 4),
        }

    @staticmethod
    def dedupe_chunks(chunks: list[dict[str, Any]]) -> list[dict[str, Any]]:
        seen: set[str] = set()
        result: list[dict[str, Any]] = []
        for chunk in chunks:
            key = f"{chunk.get('section', '')}|{chunk.get('page', 0)}|{chunk.get('content', '')[:200]}"
            if key in seen:
                continue
            seen.add(key)
            result.append(chunk)
        return result
