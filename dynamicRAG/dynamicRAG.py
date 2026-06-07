from __future__ import annotations

import json
import logging
import os
import time
import uuid
from contextvars import ContextVar
from datetime import datetime, timezone
from functools import lru_cache
from pathlib import Path
from typing import Any

from dotenv import load_dotenv
from langchain_community.document_loaders import (
    Docx2txtLoader,
    PyPDFLoader,
    TextLoader,
    UnstructuredWordDocumentLoader,
)
from langchain_community.vectorstores import FAISS
from langchain_core.documents import Document
from langchain_openai import OpenAIEmbeddings
from langchain_text_splitters import RecursiveCharacterTextSplitter
from openai import OpenAI
from werkzeug.datastructures import FileStorage
from werkzeug.utils import secure_filename

load_dotenv()

logger = logging.getLogger(__name__)

MODEL_NAME = os.getenv("DYNAMIC_RAG_MODEL", "gpt-4o-mini")
EMBEDDING_MODEL = os.getenv("DYNAMIC_RAG_EMBEDDING_MODEL", "text-embedding-3-small")
TOP_K = os.getenv("DYNAMIC_RAG_TOP_K", 4)
CHUNK_SIZE = os.getenv("DYNAMIC_RAG_CHUNK_SIZE", 1000)
CHUNK_OVERLAP = os.getenv("DYNAMIC_RAG_CHUNK_OVERLAP", 200)

MODULE_DIR = Path(__file__).resolve().parent
UPLOADS_DIR = MODULE_DIR / "uploads"

SUPPORTED_EXTENSIONS = frozenset({".pdf", ".txt", ".docx", ".doc", ".xlsx", ".xls"})

OPENAI_API_KEY = os.getenv("OPENAI_API_KEY", "")
OPENAI_CLIENT = OpenAI(api_key=OPENAI_API_KEY) if OPENAI_API_KEY else None

_trace_id_ctx: ContextVar[str | None] = ContextVar("dynamic_rag_trace_id", default=None)


def _utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _setup_openai_tracing() -> bool:
    """Enable OpenTelemetry OpenAI instrumentation when available."""
    try:
        from opentelemetry.instrumentation.openai_v2 import OpenAIInstrumentor  # pyright: ignore[reportMissingImports]

        OpenAIInstrumentor().instrument()
        logger.info("OpenAI OpenTelemetry tracing enabled for Dynamic RAG.")
        return True
    except ImportError:
        logger.info(
            "OpenAI OpenTelemetry tracing not installed; using structured trace logs only."
        )
        return False


OPENAI_TRACING_ENABLED = _setup_openai_tracing()


def _ensure_uploads_dir() -> Path:
    UPLOADS_DIR.mkdir(parents=True, exist_ok=True)
    return UPLOADS_DIR


def clear_uploads_folder() -> None:
    """Remove all files from the uploads directory."""
    uploads_dir = _ensure_uploads_dir()
    for item in uploads_dir.iterdir():
        if item.is_file():
            logger.info("Deleting previous upload: %s", item.name)
            item.unlink()


def _validate_extension(filename: str) -> str:
    suffix = Path(filename).suffix.lower()
    if suffix not in SUPPORTED_EXTENSIONS:
        supported = ", ".join(sorted(SUPPORTED_EXTENSIONS))
        raise ValueError(
            f"Unsupported file type '{suffix or '(none)'}'. "
            f"Allowed types: {supported}"
        )
    return suffix


def save_uploaded_file(file_storage: FileStorage) -> Path:
    """Replace any existing upload and save the new file under uploads/."""
    if not file_storage or not file_storage.filename:
        raise ValueError("No file was uploaded.")

    original_name = file_storage.filename
    _validate_extension(original_name)

    safe_name = secure_filename(original_name)
    if not safe_name:
        raise ValueError("Invalid filename.")

    clear_uploads_folder()
    _ensure_uploads_dir()
    destination = UPLOADS_DIR / safe_name
    file_storage.save(destination)
    invalidate_vector_db_cache()
    logger.info("Saved uploaded file: %s", destination)
    return destination


def get_uploaded_file() -> Path | None:
    """Return the single supported upload in uploads/, or None if empty."""
    uploads_dir = _ensure_uploads_dir()
    files = [
        item
        for item in uploads_dir.iterdir()
        if item.is_file() and item.suffix.lower() in SUPPORTED_EXTENSIONS
    ]
    if not files:
        return None
    if len(files) > 1:
        logger.warning(
            "Multiple files found in uploads folder; using the most recently modified."
        )
        files.sort(key=lambda path: path.stat().st_mtime, reverse=True)
    return files[0]


def get_uploaded_file_info() -> dict[str, str] | None:
    uploaded = get_uploaded_file()
    if not uploaded:
        return None
    return {
        "name": uploaded.name,
        "extension": uploaded.suffix.lower(),
    }


def _load_excel(file_path: Path) -> list[Document]:
    suffix = file_path.suffix.lower()
    documents: list[Document] = []

    if suffix == ".xlsx":
        from openpyxl import load_workbook

        workbook = load_workbook(file_path, read_only=True, data_only=True)
        try:
            for page_index, sheet in enumerate(workbook.worksheets):
                rows: list[str] = []
                for row in sheet.iter_rows(values_only=True):
                    cells = [str(cell) if cell is not None else "" for cell in row]
                    if any(cell.strip() for cell in cells):
                        rows.append("\t".join(cells))
                content = "\n".join(rows).strip()
                if content:
                    documents.append(
                        Document(
                            page_content=content,
                            metadata={
                                "source": str(file_path),
                                "page": page_index,
                                "sheet": sheet.title,
                            },
                        )
                    )
        finally:
            workbook.close()
    elif suffix == ".xls":
        import xlrd

        workbook = xlrd.open_workbook(file_path)
        for page_index in range(workbook.nsheets):
            sheet = workbook.sheet_by_index(page_index)
            rows: list[str] = []
            for row_idx in range(sheet.nrows):
                cells = [
                    str(sheet.cell_value(row_idx, col_idx))
                    for col_idx in range(sheet.ncols)
                ]
                if any(cell.strip() for cell in cells):
                    rows.append("\t".join(cells))
            content = "\n".join(rows).strip()
            if content:
                documents.append(
                    Document(
                        page_content=content,
                        metadata={
                            "source": str(file_path),
                            "page": page_index,
                            "sheet": sheet.name,
                        },
                    )
                )
    else:
        raise ValueError(f"Unsupported Excel format: {suffix}")

    if not documents:
        raise RuntimeError(f"No readable content found in Excel file: {file_path.name}")

    return documents


def load_document(file_path: Path) -> list[Document]:
    """Load a document using the appropriate loader for its file type."""
    if not file_path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")

    suffix = file_path.suffix.lower()
    if suffix not in SUPPORTED_EXTENSIONS:
        supported = ", ".join(sorted(SUPPORTED_EXTENSIONS))
        raise ValueError(
            f"Unsupported file type '{suffix}'. Allowed types: {supported}"
        )

    logger.info("Loading document: %s (type=%s)", file_path.name, suffix)

    try:
        if suffix == ".pdf":
            documents = PyPDFLoader(str(file_path)).load()
        elif suffix == ".txt":
            documents = TextLoader(str(file_path), encoding="utf-8").load()
        elif suffix == ".docx":
            documents = Docx2txtLoader(str(file_path)).load()
        elif suffix == ".doc":
            documents = UnstructuredWordDocumentLoader(str(file_path)).load()
        elif suffix in (".xlsx", ".xls"):
            documents = _load_excel(file_path)
        else:
            raise ValueError(f"Unsupported file type: {suffix}")
    except Exception as exc:
        logger.exception("Document loading failed for %s", file_path.name)
        raise RuntimeError(
            f"Failed to load document '{file_path.name}': {exc}"
        ) from exc

    if not documents:
        raise RuntimeError(f"No content extracted from document: {file_path.name}")

    logger.info(
        "Loaded document: %s (%d segment(s))",
        file_path.name,
        len(documents),
    )
    return documents


def get_document_preview_text(max_chars: int = 50_000) -> str:
    """Return plain-text preview content for the uploaded document."""
    uploaded = get_uploaded_file()
    if not uploaded:
        raise FileNotFoundError("No document has been uploaded yet.")

    suffix = uploaded.suffix.lower()
    if suffix == ".txt":
        text = uploaded.read_text(encoding="utf-8", errors="replace")
    else:
        documents = load_document(uploaded)
        text = "\n\n".join(doc.page_content for doc in documents if doc.page_content)

    text = text.strip()
    if not text:
        raise RuntimeError("Document contains no readable text.")

    if len(text) > max_chars:
        return text[:max_chars] + "\n\n… (preview truncated)"
    return text


def invalidate_vector_db_cache() -> None:
    _get_vector_db.cache_clear()


def _vector_db_cache_key() -> str:
    uploaded = get_uploaded_file()
    if not uploaded:
        return ""
    stat = uploaded.stat()
    return f"{uploaded}:{stat.st_mtime_ns}"


def _extract_token_usage(response: Any) -> dict[str, int]:
    usage = getattr(response, "usage", None)
    if not usage:
        return {"prompt_tokens": 0, "completion_tokens": 0, "total_tokens": 0}
    return {
        "prompt_tokens": getattr(usage, "prompt_tokens", 0) or 0,
        "completion_tokens": getattr(usage, "completion_tokens", 0) or 0,
        "total_tokens": getattr(usage, "total_tokens", 0) or 0,
    }


def _log_api_call(api_log: dict[str, Any]) -> None:
    logger.info("dynamic_rag_openai_call %s", json.dumps(api_log, default=str))


def _chunk_section_label(content: str, metadata: dict[str, Any]) -> str:
    if sheet := metadata.get("sheet"):
        return str(sheet)[:100]
    for line in content.splitlines():
        cleaned = line.strip().lstrip("#").strip()
        if cleaned and len(cleaned) > 3:
            return cleaned[:100]
    return "Document"


@lru_cache(maxsize=1)
def _get_vector_db(cache_key: str) -> FAISS:
    if not cache_key:
        raise FileNotFoundError(
            "No document uploaded. Please upload a PDF, TXT, Word, or Excel file first."
        )

    uploaded = get_uploaded_file()
    if not uploaded:
        raise FileNotFoundError(
            "Uploads folder is empty. Please upload a document first."
        )

    documents = load_document(uploaded)

    splitter = RecursiveCharacterTextSplitter(
        chunk_size=CHUNK_SIZE,
        chunk_overlap=CHUNK_OVERLAP,
    )
    chunks = splitter.split_documents(documents)

    embeddings = OpenAIEmbeddings(model=EMBEDDING_MODEL)
    return FAISS.from_documents(chunks, embeddings)


def _embed_query(query: str, trace_id: str) -> tuple[list[float], dict[str, Any]]:
    if not OPENAI_CLIENT:
        raise RuntimeError("OPENAI_API_KEY is missing in environment.")

    started_at = _utc_now_iso()
    t0 = time.perf_counter()

    response = OPENAI_CLIENT.embeddings.create(
        model=EMBEDDING_MODEL,
        input=query,
        extra_headers={"X-Client-Trace-Id": trace_id},
    )

    elapsed_ms = round((time.perf_counter() - t0) * 1000, 2)
    ended_at = _utc_now_iso()
    usage = _extract_token_usage(response)

    api_log = {
        "stage": "embedding",
        "model": EMBEDDING_MODEL,
        "query": query,
        "trace_id": trace_id,
        "openai_tracing_enabled": OPENAI_TRACING_ENABLED,
        "request_started_at": started_at,
        "response_received_at": ended_at,
        "elapsed_ms": elapsed_ms,
        **usage,
    }
    _log_api_call(api_log)
    return response.data[0].embedding, api_log


def _retrieve_chunks(
    query: str,
    trace_id: str,
) -> tuple[list[dict[str, Any]], dict[str, Any], dict[str, Any]]:
    vector_db = _get_vector_db(_vector_db_cache_key())
    query_vector, embed_log = _embed_query(query, trace_id)

    started_at = _utc_now_iso()
    t0 = time.perf_counter()

    docs = vector_db.similarity_search_by_vector(query_vector, k=TOP_K)

    elapsed_ms = round((time.perf_counter() - t0) * 1000, 2)
    ended_at = _utc_now_iso()

    retrieved: list[dict[str, Any]] = []
    for doc in docs:
        page_index = int(doc.metadata.get("page", 0)) + 1
        section = _chunk_section_label(doc.page_content, doc.metadata)
        retrieved.append(
            {
                "section": section,
                "page": page_index,
                "content": doc.page_content,
            }
        )

    retrieval_log = {
        "stage": "vector_retrieval",
        "query": query,
        "trace_id": trace_id,
        "openai_tracing_enabled": OPENAI_TRACING_ENABLED,
        "request_started_at": started_at,
        "response_received_at": ended_at,
        "elapsed_ms": elapsed_ms,
        "top_k": TOP_K,
        "retrieved_sections": [
            {"section": item["section"], "page": item["page"]} for item in retrieved
        ],
    }
    _log_api_call(retrieval_log)

    return retrieved, embed_log, retrieval_log


def _generate_answer(
    query: str,
    chunks: list[dict[str, Any]],
    trace_id: str,
    document_name: str,
) -> tuple[str, dict[str, Any]]:
    if not OPENAI_CLIENT:
        raise RuntimeError("OPENAI_API_KEY is missing in environment.")

    if not chunks:
        return "No relevant information found in the document.", {}

    context_parts = []
    for chunk in chunks:
        context_parts.append(
            f"""
SECTION: {chunk['section']}
PAGE: {chunk['page']}

CONTENT:
{chunk['content']}
"""
        )

    context_str = "\n\n".join(context_parts)

    prompt = f"""
Answer the question using only the provided context from the uploaded document "{document_name}".

Rules:
- Be clear and practical.
- Every document-derived claim MUST include a citation in this exact format:
  (Section: <section title>, Page: <page number>)
- Use the SECTION and PAGE values from the context blocks.
- If the context is insufficient, say so briefly.

Context:
{context_str}

Question:
{query}
"""

    started_at = _utc_now_iso()
    t0 = time.perf_counter()

    response = OPENAI_CLIENT.chat.completions.create(
        model=MODEL_NAME,
        messages=[{"role": "user", "content": prompt}],
        temperature=0.3,
        store=True,
        extra_headers={"X-Client-Trace-Id": trace_id},
        metadata={"trace_id": trace_id, "stage": "answer_generation", "workflow": "dynamic_rag"},
    )

    elapsed_ms = round((time.perf_counter() - t0) * 1000, 2)
    ended_at = _utc_now_iso()
    usage = _extract_token_usage(response)

    api_log = {
        "stage": "answer_generation",
        "model": MODEL_NAME,
        "query": query,
        "trace_id": trace_id,
        "openai_tracing_enabled": OPENAI_TRACING_ENABLED,
        "request_started_at": started_at,
        "response_received_at": ended_at,
        "elapsed_ms": elapsed_ms,
        **usage,
    }
    _log_api_call(api_log)

    answer = (response.choices[0].message.content or "").strip()
    return answer, api_log


def ask_dynamic_rag_question(query: str) -> dict[str, Any]:
    cleaned_query = (query or "").strip()
    if not cleaned_query:
        raise ValueError("Question is required.")

    if not OPENAI_API_KEY:
        raise RuntimeError("OPENAI_API_KEY is missing in environment.")

    uploaded = get_uploaded_file()
    if not uploaded:
        raise FileNotFoundError(
            "No document uploaded. Please upload a PDF, TXT, Word, or Excel file first."
        )

    trace_id = str(uuid.uuid4())
    _trace_id_ctx.set(trace_id)

    workflow_started_at = _utc_now_iso()
    t0 = time.perf_counter()

    chunks, embed_log, retrieval_log = _retrieve_chunks(cleaned_query, trace_id)
    answer, answer_log = _generate_answer(
        cleaned_query, chunks, trace_id, uploaded.name
    )

    elapsed_ms = round((time.perf_counter() - t0) * 1000, 2)
    workflow_log = {
        "stage": "workflow",
        "query": cleaned_query,
        "trace_id": trace_id,
        "openai_tracing_enabled": OPENAI_TRACING_ENABLED,
        "request_started_at": workflow_started_at,
        "response_received_at": _utc_now_iso(),
        "elapsed_ms": elapsed_ms,
        "document_path": str(uploaded),
        "document_name": uploaded.name,
    }
    _log_api_call(workflow_log)

    api_logs = [embed_log, retrieval_log, answer_log, workflow_log]

    return {
        "answer": answer,
        "trace_id": trace_id,
        "retrieved_chunks": [
            {"section": chunk["section"], "page": chunk["page"]} for chunk in chunks
        ],
        "api_logs": api_logs,
    }


_ensure_uploads_dir()
